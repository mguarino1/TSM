#Writer: Michael Guarino
#TSM Report (updated)

import easygui
import extract_msg
import re
import io
import openpyxl
import datetime
import os

#File open and message/date extraction
f = easygui.fileopenbox()
date = os.path.basename(f)[:-4]
msg = extract_msg.Message(f)
msg_report = msg.body

#Locates nodes
missedS = msg_report.find("---Online Nodes: Missed---")
missedE = msg_report.find("---Online Nodes: Failed---")

failedS = msg_report.find("---Online Nodes: Failed---")
failedE = msg_report.find("---Nodes offline, unreachable, or status unknown---")

offlineS = msg_report.find("---Nodes offline, unreachable, or status unknown---")
offlineE = msg_report.find("In addition,")

nodesMissed = msg_report[missedS:missedE]
nodesFailed = msg_report[failedS:failedE]
nodesOffline = msg_report[offlineS:offlineE]

iterMissed = io.StringIO(nodesMissed)
iterFailed = io.StringIO(nodesFailed)
iterOffline = io.StringIO(nodesOffline)

#Pulls nodes and statuses
missed = [re.split(r'\t+', line)[1].strip() for line in iterMissed if line[0]=='M']
print(missed)

failed = [re.split(r'\t+', line)[1].strip() for line in iterMissed if line[0]=='F']
print(failed)

offlineM = [re.split(r'\t+', line)[1].strip() for line in iterOffline if line[0]=='M']
offlineF = [re.split(r'\t+', line)[1].strip() for line in iterOffline if line[0]=='F']
print(offlineM)
print(offlineF)


"""
for line in iterMissed:
    if line[0] == 'M':
        temp = re.split(r'\t+', line)
        missed.append(str(temp[0])[0])

for line in iterFailed:
    if line[0] == 'F':
        temp = re.split(r'\t+', line)
        missed.append(str(temp[0])[0])    

#status = []
#node = []
#for line in iterString:
#    if line[0] == 'F' or line[0] == 'M':
#        temp = re.split(r'\t+', line)
#        status.append(str(temp[0])[0])
#        node.append(temp[1])

#Opens Excel sheet for editing
from openpyxl import load_workbook
wb = load_workbook('TSM Test.xlsx')
ws = wb['2020']

#Finding date column in sheet
colCurr = 0
dateFound = False
for col in ws.iter_cols(min_row=1, max_col=366, max_row=1):
    for cell in col:
        if cell.value != "Node Name":
            temp = str(cell.value)[:-9]
            year, month, day = temp.split('-')
            temp = year + '-' + month + '-' + day
            if str(temp) == str(date):
                colCurr = cell.column
                print("Column (" + date + ") found")
                dateFound = True

if dateFound == False:
    print("Error: Date (" + date + ") not found.")
        
#Finds node and inputs status
rowCurr = 0
cnt = 0
nodesNotFound = []
notFoundStatus = []
failed_missed_arr = []
if colCurr != 0:
    for n in node:
        nodeFound = False
        for row in ws.iter_rows(min_col=1, max_row=1000, max_col=1):
            for cell in row:
                if str(n) == str(cell.value).strip():
                    nodeFound = True
                    rowCurr = cell.row        
                    target = ws.cell(row=rowCurr, column=colCurr)
                    prev = ws.cell(row=rowCurr, column=colCurr-1)
                    if prev.value is not None:
                        prevState = str(prev.value)
                        failState, failCount = prevState.split('-')
                        if failState == status[cnt]:
                            newCount = int(failCount) + 1
                            failed_missed_arr.append(str(newCount))
                            target.value = str(str(status[cnt]) + '-' + str(newCount))
                        else:
                            target.value = status[cnt] + '-1'
                            failed_missed_arr.append('1')
                    else:
                        target.value = status[cnt] + '-1'
                        failed_missed_arr.append('1')
                    
        if nodeFound == False:
            nodesNotFound.append(n)
            notFoundStatus.append(status[cnt])
            
        cnt = cnt+1
        
print("Not found in sheet: " + str(nodesNotFound))
print("Inserting into sheet...")

#Generates text with missed/failed nodes

f = open("failuresTSM.txt", "w")
f.write(date)
f.write('\n')
index = 0
#Array Missed/Failed
arr_mf = []
#Node references the node variable from line 27 which is an array of names of nodes
for nodes in node:
    if status[index]:
        if len(node) == len(status):
            if len(failed_missed_arr) == len(node):
                f.write(nodes + ',' + status[index] + '-' + failed_missed_arr[index])
            else:
                f.write(nodes + ',' + status[index] + '-')
            f.write('\n')
            index += 1
#-----------------------------------------------------------------------
            
ins = -1
notFoundCnt = 0
for n in nodesNotFound:
    posFound = False
    for row in ws.iter_rows(min_col=1, max_row=750, max_col=1, min_row=2):
        for cell in row:
            if str(n) < str(cell.value) and posFound == False:
                ins = cell.row
                posFound = True
    ws.insert_rows(ins)
    ws.cell(row=ins, column=1).value = str(n)
    ws.cell(row=ins, column=colCurr).value = str(notFoundStatus[notFoundCnt] + "-1")
    notFoundCnt = notFoundCnt+1


print("Done")  
wb.save('TSM Test.xlsx')
    
#print(status)
#print(node)--------------------------------------------------------------------
            
ins = -1
notFoundCnt = 0
for n in nodesNotFound:
    posFound = False
    for row in ws.iter_rows(min_col=1, max_row=750, max_col=1, min_row=2):
        for cell in row:
            if str(n) < str(cell.value) and posFound == False:
                ins = cell.row
                posFound = True
    ws.insert_rows(ins)
    ws.cell(row=ins, column=1).value = str(n)
    ws.cell(row=ins, column=colCurr).value = str(notFoundStatus[notFoundCnt] + "-1")
    notFoundCnt = notFoundCnt+1


print("Done")  
wb.save('TSM Test.xlsx')
    
#print(status)
#print(node)
"""
